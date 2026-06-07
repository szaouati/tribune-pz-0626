"""
AXE 5 — Test de causalité INTERNE à la France (lève la limite « corrélation ≠ causalité »).

L'idée : à l'intérieur d'un même pays, les institutions, le niveau de richesse et le modèle
administratif sont constants. Si, d'un département à l'autre, la densité de population commande
le maillage public par habitant, alors le mécanisme « territoire dispersé -> plus de points de
service à financer » est démontré indépendamment des effets nationaux.

On croise, par département (métropole, 96 unités) :
  - densité de population        [geo.api.gouv.fr — superficie + population communales agrégées]
  - équipements publics          [INSEE Base Permanente des Équipements (BPE) via API Mélodi, 2024]
      écoles 1er degré (C107/C108/C109), collèges (C201), lycées (C301/C302/C303),
      police (A101) + gendarmerie (A104)
  - dépense communale par hab    [OFGL — base communes, dépenses totales / population, 2023]

Sortie : data/axe5/06_gradient_departements_fr.csv (+ feuille ajoutée à l'Excel de l'axe).
"""
import numpy as np
import pandas as pd
import requests
import config as C
import sources as S
import livraison as L

A = C.DATA / "axe5"; A.mkdir(parents=True, exist_ok=True)
SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "tribune-pz-0626 data collection (research)"})

# ---------------------------------------------------------------------------
# 1) Superficie + population par département (geo.api.gouv.fr, communes agrégées)
# ---------------------------------------------------------------------------
print("Superficie + population par département (geo.api.gouv.fr)…")
cache = C.RAW / "geoapi_communes.csv"
url = "https://geo.api.gouv.fr/communes"
r = SESSION.get(url, params={"fields": "surface,population,codeDepartement", "format": "json"}, timeout=120)
r.raise_for_status()
com = pd.DataFrame(r.json())
com.to_csv(cache, index=False, encoding="utf-8")
com = com.dropna(subset=["surface", "population", "codeDepartement"])
geo = (com.groupby("codeDepartement")
          .agg(superficie_km2=("surface", lambda s: s.sum() / 100.0),   # surface en hectares -> km²
               population=("population", "sum"))
          .reset_index().rename(columns={"codeDepartement": "dep"}))
S._journaliser("geoapi_communes_superficie", "Découpage administratif (geo.api.gouv.fr, données IGN/INSEE)",
               url, len(com))

# ---------------------------------------------------------------------------
# 2) Équipements publics par département (BPE INSEE, millésime 2024)
# ---------------------------------------------------------------------------
print("Équipements publics par département (BPE / Mélodi)…")
TYPES = {
    "ecoles_1d":   ["C107", "C108", "C109"],   # maternelle, primaire (regroupée), élémentaire
    "colleges":    ["C201"],
    "lycees":      ["C301", "C302", "C303"],
    "police":      ["A101"],
    "gendarmerie": ["A104"],
}
def bpe_par_dep(codes):
    tot = None
    for code in codes:
        df = S.insee_melodi("DS_BPE", serie=f"BPE_{code}",
                            FACILITY_TYPE=code, BPE_MEASURE="FACILITIES", GEO="DEP")
        df = df[df["GEO"].str.contains("-DEP-")].copy()
        df["dep"] = df["GEO"].str.split("-").str[-1]
        s = df.groupby("dep")["valeur"].sum()
        tot = s if tot is None else tot.add(s, fill_value=0)
    return tot
equip = pd.DataFrame({nom: bpe_par_dep(codes) for nom, codes in TYPES.items()})
equip = equip.fillna(0).reset_index().rename(columns={"index": "dep"})

# ---------------------------------------------------------------------------
# 3) Dépense communale par habitant, par département (OFGL, 2023)
# ---------------------------------------------------------------------------
print("Dépense communale par habitant par département (OFGL)…")
ofgl = S.opendatasoft(
    "data.ofgl.fr", "ofgl-base-communes", serie="OFGL_communes_depenses_dep",
    where='year(exer)=2023 AND agregat="Dépenses totales" AND type_de_budget="Budget principal"',
    select="sum(montant) as montant, sum(ptot_n) as pop_ofgl", group_by="dep_code")
ofgl = ofgl.rename(columns={"dep_code": "dep"})
ofgl["eur_hab_communal"] = (ofgl["montant"] / ofgl["pop_ofgl"]).round(0)

# ---------------------------------------------------------------------------
# 4) Assemblage (métropole : on écarte les DOM 97x, atypiques) + indicateurs
# ---------------------------------------------------------------------------
df = (geo.merge(equip, on="dep").merge(ofgl[["dep", "eur_hab_communal"]], on="dep"))
df = df[~df["dep"].str.startswith("97")].copy()           # métropole + Corse
df["densite_hab_km2"] = (df["population"] / df["superficie_km2"]).round(1)
df["ecoles_pour_10k"] = (df["ecoles_1d"] / df["population"] * 1e4).round(2)
df["securite_pour_10k"] = ((df["police"] + df["gendarmerie"]) / df["population"] * 1e4).round(2)
df["second_degre_pour_10k"] = ((df["colleges"] + df["lycees"]) / df["population"] * 1e4).round(2)
for c in ["ecoles_1d", "colleges", "lycees", "police", "gendarmerie"]:
    df[c] = df[c].astype(int)
df = df.sort_values("densite_hab_km2").reset_index(drop=True)
df = df[["dep", "densite_hab_km2", "population", "superficie_km2",
         "ecoles_1d", "ecoles_pour_10k", "police", "gendarmerie", "securite_pour_10k",
         "colleges", "lycees", "second_degre_pour_10k", "eur_hab_communal"]]
L.ecrire_csv(df, A / "06_gradient_departements_fr.csv")

# ---------------------------------------------------------------------------
# 5) Corrélations (sur densité en log10) — la mesure du gradient
# ---------------------------------------------------------------------------
def corr(y):
    x = np.log10(df["densite_hab_km2"].values)
    m = np.isfinite(x) & np.isfinite(y.values)
    return float(np.corrcoef(x[m], y.values[m])[0, 1])
CORR = {
    "ecoles_pour_10k": corr(df["ecoles_pour_10k"]),
    "securite_pour_10k": corr(df["securite_pour_10k"]),
    "second_degre_pour_10k": corr(df["second_degre_pour_10k"]),
    "eur_hab_communal": corr(df["eur_hab_communal"]),
}
corr_df = pd.DataFrame([{"indicateur": k, "correlation_log_densite": round(v, 3)} for k, v in CORR.items()])
L.ecrire_csv(corr_df, A / "06b_gradient_correlations.csv")

# ---------------------------------------------------------------------------
# 6) Ajout d'une feuille à l'Excel de l'axe (openpyxl, sans toucher au reste)
# ---------------------------------------------------------------------------
try:
    from openpyxl import load_workbook
    xlsx = A / "axe5_territoire_depense.xlsx"
    if xlsx.exists():
        wbk = load_workbook(xlsx)
        for nom in ("Gradient_departements_FR", "Gradient_correlations"):
            if nom in wbk.sheetnames:
                del wbk[nom]
        ws = wbk.create_sheet("Gradient_departements_FR")
        ws.append(list(df.columns))
        for _, row in df.iterrows():
            ws.append([row[c] for c in df.columns])
        ws2 = wbk.create_sheet("Gradient_correlations")
        ws2.append(["indicateur", "corrélation avec log10(densité)"])
        for _, row in corr_df.iterrows():
            ws2.append([row["indicateur"], row["correlation_log_densite"]])
        wbk.save(xlsx)
        print("  ✓ feuilles 'Gradient_*' ajoutées à l'Excel")
except Exception as e:
    print(f"  (Excel non mis à jour : {e})")

print("\nRécap gradient interne France (96 départements métropolitains) :")
for k, v in CORR.items():
    print(f"  corr(log densité, {k:>22}) = {v:+.2f}")
print("\n5 départements les moins denses :")
print(df.head(5)[["dep", "densite_hab_km2", "ecoles_pour_10k", "securite_pour_10k", "eur_hab_communal"]].to_string(index=False))
print("\n5 départements les plus denses :")
print(df.tail(5)[["dep", "densite_hab_km2", "ecoles_pour_10k", "securite_pour_10k", "eur_hab_communal"]].to_string(index=False))
